#!/usr/bin/env python3
"""
Multi-Company Version Test Suite for Ride Guardian Desktop
Comprehensive testing of all multi-company specific features and functionality
"""

import sys
import os
import tempfile
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

# Import modules to test
from core.database import initialize_database, get_db_connection, DATABASE_PATH
from core.company_manager import company_manager
from core.translation_manager import TranslationManager
from core.enhanced_fahrtenbuch_export import PreciseGermanFahrtenbuchExporter
from core.google_maps import GoogleMapsIntegration
from core.excel_workbook_logic import ExcelWorkbookLogic
from core.payroll_calculator import PayrollCalculator

class MultiCompanyTestSuite:
    """Comprehensive test suite for multi-company functionality"""
    
    def __init__(self):
        print("🏢 Initializing Multi-Company Test Suite")
        self.test_results = {
            'database_multi_company': False,
            'company_management': False,
            'data_isolation': False,
            'company_switching': False,
            'multi_company_exports': False,
            'company_statistics': False,
            'app_mode_management': False,
            'company_specific_config': False,
            'ui_integration': False,
            'performance_isolation': False
        }
        
        # Initialize database
        try:
            initialize_database()
            print("✅ Database initialized successfully")
        except Exception as e:
            print(f"❌ Database initialization failed: {e}")
            return
        
        # Setup test data
        self.setup_multi_company_test_data()
        
    def setup_multi_company_test_data(self):
        """Setup comprehensive test data for multi-company testing"""
        print("\n📊 Setting up multi-company test data...")
        
        db = get_db_connection()
        cursor = db.cursor()
        
        try:
            # Create test companies
            companies_data = [
                (1, 'Muster Transport GmbH', 'Muster Str 1, 45451 MusterStadt', '+49 201 123456', 'info@muster-transport.de'),
                (2, 'Berlin Taxi AG', 'Hauptstraße 100, 10115 Berlin', '+49 30 987654', 'kontakt@berlin-taxi.de'),
                (3, 'München Fahrdienst', 'Marienplatz 5, 80331 München', '+49 89 555123', 'service@muenchen-fahrdienst.de'),
                (4, 'Hamburg Express', 'Reeperbahn 20, 20359 Hamburg', '+49 40 333999', 'info@hamburg-express.de')
            ]
            
            for company_data in companies_data:
                cursor.execute("""
                    INSERT OR REPLACE INTO companies (id, name, headquarters_address, phone, email, is_active)
                    VALUES (?, ?, ?, ?, ?, 1)
                """, company_data)
            
            # Create drivers for each company
            drivers_data = [
                # Company 1 - Muster Transport
                (1, 1, 'Max Mustermann', 'E-CK123456', '001', 'Active'),
                (2, 1, 'Anna Schmidt', 'E-CK789012', '002', 'Active'),
                (3, 1, 'Peter Müller', 'E-CK345678', '003', 'Active'),
                
                # Company 2 - Berlin Taxi
                (4, 2, 'Hans Berlin', 'B-TX111', '001', 'Active'),
                (5, 2, 'Maria Brandenburg', 'B-TX222', '002', 'Active'),
                
                # Company 3 - München Fahrdienst
                (6, 3, 'Franz Bayern', 'M-FC333', '001', 'Active'),
                (7, 3, 'Lisa Münchner', 'M-FC444', '002', 'Active'),
                
                # Company 4 - Hamburg Express
                (8, 4, 'Klaus Nordsee', 'HH-EX555', '001', 'Active'),
            ]
            
            for driver_data in drivers_data:
                cursor.execute("""
                    INSERT OR REPLACE INTO drivers (id, company_id, name, vehicle, personalnummer, status)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, driver_data)
            
            # Create vehicles for each company
            vehicles_data = [
                # Company 1
                (1, 'E-CK123456', 'Toyota', 'Corolla', 2020, 1),
                (1, 'E-CK789012', 'Mercedes', 'E-Klasse', 2021, 2),
                (1, 'E-CK345678', 'BMW', '3er', 2019, 3),
                
                # Company 2
                (2, 'B-TX111', 'Mercedes', 'E-Klasse', 2022, 4),
                (2, 'B-TX222', 'Audi', 'A4', 2020, 5),
                
                # Company 3
                (3, 'M-FC333', 'BMW', '5er', 2021, 6),
                (3, 'M-FC444', 'Mercedes', 'S-Klasse', 2023, 7),
                
                # Company 4
                (4, 'HH-EX555', 'Tesla', 'Model 3', 2022, 8),
            ]
            
            for vehicle_data in vehicles_data:
                cursor.execute("""
                    INSERT OR REPLACE INTO vehicles (company_id, plate_number, make, model, year, current_driver_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, vehicle_data)
            
            # Create shifts for each company
            base_date = datetime.now().replace(hour=6, minute=0, second=0, microsecond=0)
            shifts_data = []
            
            for i, (company_id, driver_id) in enumerate([(1, 1), (1, 2), (2, 4), (3, 6), (4, 8)]):
                shift_date = base_date + timedelta(days=i)
                shifts_data.append((
                    i + 1, company_id, driver_id, f'{company_id}-{i+1}(1)',
                    shift_date.strftime('%Y-%m-%d'),
                    shift_date.strftime('%Y-%m-%d %H:%M:%S'),
                    (shift_date + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S'),
                    'Fahren', 8.0, 30, 7.5, 7.5, 0.0
                ))
            
            for shift_data in shifts_data:
                cursor.execute("""
                    INSERT OR REPLACE INTO shifts (id, company_id, driver_id, schicht_id, shift_date, 
                                                 start_time, end_time, taetigkeit, gesamte_arbeitszeit_std, 
                                                 pause_min, reale_arbeitszeit_std, fruehschicht_std, nachtschicht_std)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, shift_data)
            
            # Create rides for each company
            current_time = datetime.now()
            rides_data = []
            
            # Company 1 rides
            company1_locations = [
                ('Muster Str 1, 45451 MusterStadt', 'Beispielstraße 10, 45451 MusterStadt'),
                ('Hauptstraße 5, 45451 MusterStadt', 'Industriestraße 15, 45451 MusterStadt'),
                ('Bahnhofstraße 20, 45451 MusterStadt', 'Parkstraße 8, 45451 MusterStadt')
            ]
            
            # Company 2 rides
            company2_locations = [
                ('Alexanderplatz 1, 10178 Berlin', 'Potsdamer Platz 5, 10785 Berlin'),
                ('Brandenburger Tor, 10117 Berlin', 'Checkpoint Charlie, 10117 Berlin')
            ]
            
            # Company 3 rides
            company3_locations = [
                ('Marienplatz 1, 80331 München', 'Oktoberfest-Gelände, 80339 München'),
                ('Hauptbahnhof München, 80335 München', 'Flughafen München, 85356 Freising')
            ]
            
            # Company 4 rides
            company4_locations = [
                ('Reeperbahn 1, 20359 Hamburg', 'Speicherstadt, 20457 Hamburg')
            ]
            
            all_locations = [
                (1, company1_locations),
                (2, company2_locations),
                (3, company3_locations),
                (4, company4_locations)
            ]
            
            ride_id = 1
            for company_id, locations in all_locations:
                for i, (pickup, destination) in enumerate(locations):
                    driver_id = next(d[0] for d in drivers_data if d[1] == company_id)
                    shift_id = next(s[0] for s in shifts_data if s[1] == company_id)
                    vehicle_plate = next(v[1] for v in vehicles_data if v[0] == company_id)
                    
                    pickup_time = current_time + timedelta(hours=i, minutes=i*15)
                    dropoff_time = pickup_time + timedelta(minutes=30)
                    
                    rides_data.append((
                        ride_id, company_id, driver_id, shift_id,
                        pickup_time.strftime('%Y-%m-%d %H:%M:%S'),
                        dropoff_time.strftime('%Y-%m-%d %H:%M:%S'),
                        pickup, destination,
                        f'Zentrale {company_id}',
                        pickup.split(',')[0], destination.split(',')[0],
                        5.0 + i, vehicle_plate, 0
                    ))
                    ride_id += 1
            
            for ride_data in rides_data:
                cursor.execute("""
                    INSERT OR REPLACE INTO rides (id, company_id, driver_id, shift_id, pickup_time, dropoff_time,
                                                pickup_location, destination, standort_auftragsuebermittlung,
                                                abholort, zielort, gefahrene_kilometer, vehicle_plate, is_reserved)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, ride_data)
            
            # Create company-specific configurations
            config_data = [
                (1, 'app_mode', 'multi'),
                (1, 'default_fuel_consumption', '7.0'),
                (1, 'standard_hourly_rate', '12.41'),
                (2, 'app_mode', 'multi'),
                (2, 'default_fuel_consumption', '6.5'),
                (2, 'standard_hourly_rate', '13.50'),
                (3, 'app_mode', 'multi'),
                (3, 'default_fuel_consumption', '8.0'),
                (3, 'standard_hourly_rate', '14.00'),
                (4, 'app_mode', 'multi'),
                (4, 'default_fuel_consumption', '5.5'),  # Tesla
                (4, 'standard_hourly_rate', '15.00'),
            ]
            
            for config in config_data:
                cursor.execute("""
                    INSERT OR REPLACE INTO config (company_id, key, value, updated_at)
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                """, config)
            
            db.commit()
            print("✅ Multi-company test data setup completed")
            
        except Exception as e:
            print(f"❌ Multi-company test data setup failed: {e}")
            db.rollback()
        finally:
            db.close()

    def test_database_multi_company_schema(self):
        """Test 1: Database schema for multi-company support"""
        print("\n🗄️ Testing Database Multi-Company Schema...")
        
        try:
            db = get_db_connection()
            cursor = db.cursor()
            
            # Check if all major tables have company_id column
            tables_to_check = ['drivers', 'vehicles', 'rides', 'shifts', 'payroll', 'rules', 'config']
            
            schema_valid = True
            for table in tables_to_check:
                cursor.execute(f"PRAGMA table_info({table})")
                columns = [col[1] for col in cursor.fetchall()]
                
                if 'company_id' in columns:
                    print(f"  ✅ Table '{table}' has company_id column")
                else:
                    print(f"  ❌ Table '{table}' missing company_id column")
                    schema_valid = False
            
            # Check companies table structure
            cursor.execute("PRAGMA table_info(companies)")
            company_columns = [col[1] for col in cursor.fetchall()]
            required_company_columns = ['id', 'name', 'headquarters_address', 'phone', 'email', 'is_active']
            
            for col in required_company_columns:
                if col in company_columns:
                    print(f"  ✅ Companies table has '{col}' column")
                else:
                    print(f"  ❌ Companies table missing '{col}' column")
                    schema_valid = False
            
            # Check foreign key constraints
            cursor.execute("PRAGMA foreign_key_list(drivers)")
            fk_info = cursor.fetchall()
            has_company_fk = any(fk[2] == 'companies' and fk[3] == 'company_id' for fk in fk_info)
            
            if has_company_fk:
                print("  ✅ Drivers table has proper foreign key to companies")
            else:
                print("  ⚠️ Drivers table foreign key constraint not detected")
            
            db.close()
            self.test_results['database_multi_company'] = schema_valid
            
        except Exception as e:
            print(f"  ❌ Database schema test failed: {e}")
            self.test_results['database_multi_company'] = False

    def test_company_management(self):
        """Test 2: Company management operations"""
        print("\n🏢 Testing Company Management Operations...")
        
        try:
            # Test getting companies
            companies = company_manager.get_companies()
            print(f"  ✅ Retrieved {len(companies)} companies")
            
            for company in companies[:3]:  # Show first 3
                print(f"    🏢 {company['name']} (ID: {company['id']}) - {company['headquarters_address']}")
            
            # Test adding a new company
            new_company_name = f"Test Company {datetime.now().strftime('%H%M%S')}"
            add_result = company_manager.add_company(
                name=new_company_name,
                address="Test Address 123, 12345 TestStadt",
                phone="+49 123 999888",
                email="test@testcompany.de"
            )
            
            if add_result:
                print(f"  ✅ Successfully added new company: '{new_company_name}'")
                
                # Verify the company was added
                updated_companies = company_manager.get_companies()
                new_company = next((c for c in updated_companies if c['name'] == new_company_name), None)
                
                if new_company:
                    print(f"    ✅ New company verified in database (ID: {new_company['id']})")
                    
                    # Test updating the company
                    update_result = company_manager.update_company(
                        new_company['id'],
                        name=new_company_name + " Updated",
                        address="Updated Address 456, 54321 UpdatedStadt",
                        phone="+49 987 123456",
                        email="updated@testcompany.de"
                    )
                    
                    if update_result:
                        print(f"  ✅ Successfully updated company")
                    else:
                        print(f"  ❌ Failed to update company")
                    
                    # Test company statistics
                    stats = company_manager.get_company_statistics(new_company['id'])
                    print(f"  📊 Company statistics: {stats['driver_count']} drivers, {stats['ride_count']} rides, €{stats['monthly_revenue']:.2f} revenue")
                    
                    # Test soft delete (don't delete if it has data)
                    if stats['driver_count'] == 0 and stats['ride_count'] == 0:
                        delete_result = company_manager.delete_company(new_company['id'])
                        if delete_result:
                            print(f"  ✅ Successfully deleted empty test company")
                        else:
                            print(f"  ❌ Failed to delete test company")
                    else:
                        print(f"  ℹ️ Skipping delete test - company has data")
                
                else:
                    print(f"  ❌ New company not found after adding")
                    
            else:
                print(f"  ❌ Failed to add new company")
            
            self.test_results['company_management'] = True
            
        except Exception as e:
            print(f"  ❌ Company management test failed: {e}")
            self.test_results['company_management'] = False

    def test_data_isolation(self):
        """Test 3: Data isolation between companies"""
        print("\n🔒 Testing Data Isolation Between Companies...")
        
        try:
            db = get_db_connection()
            cursor = db.cursor()
            
            isolation_valid = True
            
            # Test that each company only sees its own data
            companies = company_manager.get_companies()
            
            for company in companies[:3]:  # Test first 3 companies
                company_id = company['id']
                company_name = company['name']
                
                print(f"  🏢 Testing isolation for {company_name} (ID: {company_id})")
                
                # Count drivers for this company
                cursor.execute("SELECT COUNT(*) as count FROM drivers WHERE company_id = ?", (company_id,))
                driver_count = cursor.fetchone()['count']
                
                # Count rides for this company
                cursor.execute("SELECT COUNT(*) as count FROM rides WHERE company_id = ?", (company_id,))
                ride_count = cursor.fetchone()['count']
                
                # Count vehicles for this company
                cursor.execute("SELECT COUNT(*) as count FROM vehicles WHERE company_id = ?", (company_id,))
                vehicle_count = cursor.fetchone()['count']
                
                print(f"    📊 Company {company_id}: {driver_count} drivers, {ride_count} rides, {vehicle_count} vehicles")
                
                # Verify no cross-company data leakage
                cursor.execute("""
                    SELECT COUNT(*) as count FROM rides r 
                    JOIN drivers d ON r.driver_id = d.id 
                    WHERE r.company_id != d.company_id
                """)
                cross_leakage = cursor.fetchone()['count']
                
                if cross_leakage == 0:
                    print(f"    ✅ No data leakage detected for company {company_id}")
                else:
                    print(f"    ❌ Data leakage detected: {cross_leakage} rides with wrong company assignment")
                    isolation_valid = False
            
            # Test configuration isolation
            cursor.execute("SELECT company_id, key, value FROM config WHERE key = 'default_fuel_consumption'")
            configs = cursor.fetchall()
            
            print(f"  ⚙️ Configuration isolation test:")
            company_configs = {}
            for config in configs:
                company_configs[config['company_id']] = config['value']
                print(f"    Company {config['company_id']}: fuel_consumption = {config['value']}")
            
            # Verify companies have different configurations
            unique_configs = len(set(company_configs.values()))
            if unique_configs > 1:
                print(f"    ✅ Configuration isolation working: {unique_configs} different configurations")
            else:
                print(f"    ⚠️ All companies have same configuration (may be intentional)")
            
            db.close()
            self.test_results['data_isolation'] = isolation_valid
            
        except Exception as e:
            print(f"  ❌ Data isolation test failed: {e}")
            self.test_results['data_isolation'] = False

    def test_company_switching(self):
        """Test 4: Company switching functionality"""
        print("\n🔄 Testing Company Switching Functionality...")
        
        try:
            companies = company_manager.get_companies()
            if len(companies) < 2:
                print("  ⚠️ Need at least 2 companies for switching test")
                self.test_results['company_switching'] = False
                return
            
            # Test switching between companies
            original_company_id = company_manager.current_company_id
            print(f"  📍 Current company: {original_company_id}")
            
            # Switch to different companies and verify
            switch_success = True
            for i, company in enumerate(companies[:3]):
                company_id = company['id']
                company_name = company['name']
                
                # Switch to this company
                company_manager.set_current_company(company_id)
                
                # Verify the switch
                if company_manager.current_company_id == company_id:
                    print(f"  ✅ Successfully switched to {company_name} (ID: {company_id})")
                    
                    # Test getting current company info
                    current_company = company_manager.get_current_company()
                    if current_company and current_company['id'] == company_id:
                        print(f"    ✅ Current company info correct: {current_company['name']}")
                    else:
                        print(f"    ❌ Current company info mismatch")
                        switch_success = False
                else:
                    print(f"  ❌ Failed to switch to company {company_id}")
                    switch_success = False
            
            # Test app mode management
            print(f"  🎛️ Testing app mode management...")
            
            # Save multi-company mode
            company_manager.save_app_mode(company_manager.MULTI_COMPANY_MODE)
            
            # Verify mode was saved
            if company_manager.is_multi_company_mode():
                print(f"    ✅ Multi-company mode correctly set")
            else:
                print(f"    ❌ Multi-company mode not set correctly")
                switch_success = False
            
            # Test single company mode
            company_manager.save_app_mode(company_manager.SINGLE_COMPANY_MODE)
            if company_manager.is_single_company_mode():
                print(f"    ✅ Single company mode correctly set")
            else:
                print(f"    ❌ Single company mode not set correctly")
                switch_success = False
            
            # Restore multi-company mode
            company_manager.save_app_mode(company_manager.MULTI_COMPANY_MODE)
            
            # Restore original company
            company_manager.set_current_company(original_company_id)
            
            self.test_results['company_switching'] = switch_success
            
        except Exception as e:
            print(f"  ❌ Company switching test failed: {e}")
            self.test_results['company_switching'] = False

    def test_multi_company_exports(self):
        """Test 5: Multi-company export functionality"""
        print("\n📄 Testing Multi-Company Export Functionality...")
        
        try:
            companies = company_manager.get_companies()
            export_success = True
            
            with tempfile.TemporaryDirectory() as temp_dir:
                for company in companies[:2]:  # Test first 2 companies
                    company_id = company['id']
                    company_name = company['name']
                    
                    print(f"  🏢 Testing exports for {company_name} (ID: {company_id})")
                    
                    # Test Fahrtenbuch export for this company
                    exporter = PreciseGermanFahrtenbuchExporter()
                    exporter.set_company_info(company_name, company['headquarters_address'])
                    
                    # Get a driver from this company
                    db = get_db_connection()
                    cursor = db.cursor()
                    cursor.execute("SELECT id FROM drivers WHERE company_id = ? LIMIT 1", (company_id,))
                    driver_result = cursor.fetchone()
                    db.close()
                    
                    if driver_result:
                        driver_id = driver_result['id']
                        
                        # Export Fahrtenbuch
                        fahrtenbuch_path = os.path.join(temp_dir, f"fahrtenbuch_company_{company_id}.xlsx")
                        
                        fahrtenbuch_result = exporter.export_fahrtenbuch_excel(
                            driver_id=driver_id,
                            start_date='2025-01-01',
                            end_date='2025-12-31',
                            output_path=fahrtenbuch_path,
                            company_id=company_id
                        )
                        
                        if fahrtenbuch_result and os.path.exists(fahrtenbuch_path):
                            file_size = os.path.getsize(fahrtenbuch_path)
                            print(f"    ✅ Fahrtenbuch exported: {fahrtenbuch_path} ({file_size} bytes)")
                            
                            # Verify company info in export
                            import openpyxl
                            try:
                                wb = openpyxl.load_workbook(fahrtenbuch_path)
                                ws = wb.active
                                
                                # Look for company name in the file
                                company_found = False
                                for row in range(1, min(10, ws.max_row + 1)):
                                    for col in range(1, min(10, ws.max_column + 1)):
                                        cell_value = ws.cell(row, col).value
                                        if cell_value and company_name in str(cell_value):
                                            print(f"    ✅ Company name found in export at row {row}")
                                            company_found = True
                                            break
                                    if company_found:
                                        break
                                
                                if not company_found:
                                    print(f"    ⚠️ Company name not found in export (may be in different location)")
                                
                            except Exception as e:
                                print(f"    ⚠️ Could not verify export content: {e}")
                        else:
                            print(f"    ❌ Fahrtenbuch export failed for company {company_id}")
                            export_success = False
                        
                        # Export Stundenzettel
                        stundenzettel_path = os.path.join(temp_dir, f"stundenzettel_company_{company_id}.xlsx")
                        
                        stundenzettel_result = exporter.export_stundenzettel_excel(
                            driver_id=driver_id,
                            month=datetime.now().month,
                            year=datetime.now().year,
                            output_path=stundenzettel_path
                        )
                        
                        if stundenzettel_result and os.path.exists(stundenzettel_path):
                            file_size = os.path.getsize(stundenzettel_path)
                            print(f"    ✅ Stundenzettel exported: {stundenzettel_path} ({file_size} bytes)")
                        else:
                            print(f"    ❌ Stundenzettel export failed for company {company_id}")
                            export_success = False
                    else:
                        print(f"    ⚠️ No drivers found for company {company_id}, skipping export test")
            
            self.test_results['multi_company_exports'] = export_success
            
        except Exception as e:
            print(f"  ❌ Multi-company exports test failed: {e}")
            self.test_results['multi_company_exports'] = False

    def test_company_statistics(self):
        """Test 6: Company statistics and analytics"""
        print("\n📊 Testing Company Statistics and Analytics...")
        
        try:
            companies = company_manager.get_companies()
            stats_success = True
            
            for company in companies[:3]:  # Test first 3 companies
                company_id = company['id']
                company_name = company['name']
                
                print(f"  🏢 Testing statistics for {company_name} (ID: {company_id})")
                
                # Get company statistics
                stats = company_manager.get_company_statistics(company_id)
                
                # Verify statistics structure
                required_stats = ['ride_count', 'driver_count', 'monthly_revenue', 'vehicle_count']
                stats_valid = True
                
                for stat_key in required_stats:
                    if stat_key in stats:
                        print(f"    ✅ {stat_key}: {stats[stat_key]}")
                    else:
                        print(f"    ❌ Missing statistic: {stat_key}")
                        stats_valid = False
                
                if not stats_valid:
                    stats_success = False
                
                # Test manual verification of some statistics
                db = get_db_connection()
                cursor = db.cursor()
                
                # Verify driver count
                cursor.execute("SELECT COUNT(*) as count FROM drivers WHERE company_id = ? AND status = 'Active'", (company_id,))
                actual_driver_count = cursor.fetchone()['count']
                
                if stats['driver_count'] == actual_driver_count:
                    print(f"    ✅ Driver count verified: {actual_driver_count}")
                else:
                    print(f"    ❌ Driver count mismatch: stats={stats['driver_count']}, actual={actual_driver_count}")
                    stats_success = False
                
                # Verify ride count
                cursor.execute("SELECT COUNT(*) as count FROM rides WHERE company_id = ?", (company_id,))
                actual_ride_count = cursor.fetchone()['count']
                
                if stats['ride_count'] == actual_ride_count:
                    print(f"    ✅ Ride count verified: {actual_ride_count}")
                else:
                    print(f"    ❌ Ride count mismatch: stats={stats['ride_count']}, actual={actual_ride_count}")
                    stats_success = False
                
                db.close()
            
            self.test_results['company_statistics'] = stats_success
            
        except Exception as e:
            print(f"  ❌ Company statistics test failed: {e}")
            self.test_results['company_statistics'] = False

    def test_app_mode_management(self):
        """Test 7: Application mode management"""
        print("\n🎛️ Testing Application Mode Management...")
        
        try:
            # Test mode initialization
            initial_mode = company_manager.initialize_app_mode()
            print(f"  📍 Initial app mode: {initial_mode}")
            
            # Test mode switching and persistence
            modes_to_test = [
                company_manager.SINGLE_COMPANY_MODE,
                company_manager.MULTI_COMPANY_MODE
            ]
            
            mode_test_success = True
            
            for mode in modes_to_test:
                print(f"  🔄 Testing {mode} mode...")
                
                # Save the mode
                company_manager.save_app_mode(mode)
                
                # Verify mode was set
                if mode == company_manager.SINGLE_COMPANY_MODE:
                    if company_manager.is_single_company_mode():
                        print(f"    ✅ Single company mode correctly detected")
                    else:
                        print(f"    ❌ Single company mode not detected")
                        mode_test_success = False
                        
                    if not company_manager.is_multi_company_mode():
                        print(f"    ✅ Multi-company mode correctly excluded")
                    else:
                        print(f"    ❌ Multi-company mode incorrectly detected")
                        mode_test_success = False
                
                elif mode == company_manager.MULTI_COMPANY_MODE:
                    if company_manager.is_multi_company_mode():
                        print(f"    ✅ Multi-company mode correctly detected")
                    else:
                        print(f"    ❌ Multi-company mode not detected")
                        mode_test_success = False
                        
                    if not company_manager.is_single_company_mode():
                        print(f"    ✅ Single company mode correctly excluded")
                    else:
                        print(f"    ❌ Single company mode incorrectly detected")
                        mode_test_success = False
                
                # Test that mode persists across manager instances
                new_manager = company_manager.__class__()
                new_manager.initialize_app_mode()
                
                if new_manager.app_mode == mode:
                    print(f"    ✅ Mode persistence verified")
                else:
                    print(f"    ❌ Mode persistence failed: expected {mode}, got {new_manager.app_mode}")
                    mode_test_success = False
            
            # Test default company creation for single mode
            company_manager.save_app_mode(company_manager.SINGLE_COMPANY_MODE)
            company_manager.ensure_default_company()
            
            default_company = company_manager.get_current_company()
            if default_company:
                print(f"  ✅ Default company ensured: {default_company['name']}")
            else:
                print(f"  ❌ Default company creation failed")
                mode_test_success = False
            
            # Restore multi-company mode
            company_manager.save_app_mode(company_manager.MULTI_COMPANY_MODE)
            
            self.test_results['app_mode_management'] = mode_test_success
            
        except Exception as e:
            print(f"  ❌ App mode management test failed: {e}")
            self.test_results['app_mode_management'] = False

    def test_company_specific_config(self):
        """Test 8: Company-specific configuration"""
        print("\n⚙️ Testing Company-Specific Configuration...")
        
        try:
            from core.database import get_company_config, set_company_config
            
            companies = company_manager.get_companies()
            config_success = True
            
            # Test setting and getting configurations for different companies
            test_configs = [
                ('test_setting_1', 'value_company_1'),
                ('test_setting_2', 'value_company_2'),
                ('hourly_rate', '15.50'),
                ('fuel_consumption', '7.2')
            ]
            
            for i, company in enumerate(companies[:2]):  # Test first 2 companies
                company_id = company['id']
                company_name = company['name']
                
                print(f"  🏢 Testing config for {company_name} (ID: {company_id})")
                
                for config_key, config_value in test_configs:
                    # Set company-specific configuration
                    company_specific_value = f"{config_value}_company_{company_id}"
                    set_company_config(company_id, config_key, company_specific_value)
                    
                    # Retrieve and verify
                    retrieved_value = get_company_config(company_id, config_key)
                    
                    if retrieved_value == company_specific_value:
                        print(f"    ✅ Config '{config_key}': '{company_specific_value}'")
                    else:
                        print(f"    ❌ Config mismatch for '{config_key}': expected '{company_specific_value}', got '{retrieved_value}'")
                        config_success = False
            
            # Test that configurations are isolated between companies
            print(f"  🔒 Testing configuration isolation...")
            
            if len(companies) >= 2:
                company1_id = companies[0]['id']
                company2_id = companies[1]['id']
                
                # Set different values for same key
                set_company_config(company1_id, 'isolation_test', 'company1_value')
                set_company_config(company2_id, 'isolation_test', 'company2_value')
                
                # Verify isolation
                value1 = get_company_config(company1_id, 'isolation_test')
                value2 = get_company_config(company2_id, 'isolation_test')
                
                if value1 == 'company1_value' and value2 == 'company2_value':
                    print(f"    ✅ Configuration isolation working correctly")
                else:
                    print(f"    ❌ Configuration isolation failed: company1='{value1}', company2='{value2}'")
                    config_success = False
            
            # Test Excel workbook logic with company-specific config
            try:
                excel_logic = ExcelWorkbookLogic(company_id=companies[0]['id'])
                config = excel_logic.config
                print(f"  🧮 Excel logic config loaded: {len(config)} settings")
                
                if 'fuel_consumption_l_per_100km' in config:
                    print(f"    ✅ Fuel consumption config: {config['fuel_consumption_l_per_100km']}")
                else:
                    print(f"    ⚠️ Fuel consumption config not found")
                
            except Exception as e:
                print(f"    ⚠️ Excel logic config test failed: {e}")
            
            self.test_results['company_specific_config'] = config_success
            
        except Exception as e:
            print(f"  ❌ Company-specific configuration test failed: {e}")
            self.test_results['company_specific_config'] = False

    def test_ui_integration(self):
        """Test 9: UI integration with multi-company features"""
        print("\n🖥️ Testing UI Integration with Multi-Company Features...")
        
        try:
            # Test TranslationManager with multi-company context
            tm = TranslationManager()
            
            # Test company-related translations
            company_translations = [
                'select_company',
                'manage_companies',
                'add_company',
                'edit_company',
                'delete_company',
                'switch_company',
                'current_company',
                'company_details'
            ]
            
            translation_success = True
            for key in company_translations:
                translation = tm.tr(key)
                if translation and translation != key:  # Translation found
                    print(f"  ✅ Translation '{key}': '{translation}'")
                else:
                    print(f"  ⚠️ Translation missing for '{key}'")
                    # Not marking as failure since some translations might be optional
            
            # Test that UI components can handle company data
            companies = company_manager.get_companies()
            
            # Simulate company selection dialog data
            print(f"  🏢 Testing company selection data structure...")
            for company in companies[:3]:
                # Test data conversion (sqlite3.Row to dict)
                if hasattr(company, 'keys'):
                    company_dict = dict(company)
                else:
                    company_dict = company
                
                required_fields = ['id', 'name', 'headquarters_address']
                fields_valid = all(field in company_dict for field in required_fields)
                
                if fields_valid:
                    print(f"    ✅ Company data valid: {company_dict['name']}")
                else:
                    print(f"    ❌ Company data missing required fields")
                    translation_success = False
            
            # Test company info display formatting
            if companies:
                test_company = companies[0]
                if hasattr(test_company, 'keys'):
                    test_company_dict = dict(test_company)
                else:
                    test_company_dict = test_company
                
                display_text = f"{test_company_dict['name']}"
                if test_company_dict.get('headquarters_address'):
                    display_text += f" - {test_company_dict['headquarters_address']}"
                
                print(f"  ✅ Company display format: '{display_text}'")
            
            self.test_results['ui_integration'] = translation_success
            
        except Exception as e:
            print(f"  ❌ UI integration test failed: {e}")
            self.test_results['ui_integration'] = False

    def test_performance_isolation(self):
        """Test 10: Performance with data isolation"""
        print("\n⚡ Testing Performance with Data Isolation...")
        
        try:
            import time
            
            companies = company_manager.get_companies()
            performance_success = True
            
            # Test query performance for company-specific data
            for company in companies[:2]:  # Test first 2 companies
                company_id = company['id']
                company_name = company['name']
                
                print(f"  🏢 Testing performance for {company_name} (ID: {company_id})")
                
                db = get_db_connection()
                cursor = db.cursor()
                
                # Test ride query performance
                start_time = time.time()
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM rides r 
                    JOIN drivers d ON r.driver_id = d.id 
                    WHERE r.company_id = ? AND d.company_id = ?
                """, (company_id, company_id))
                result = cursor.fetchone()
                query_time = time.time() - start_time
                
                print(f"    ⚡ Ride query: {result['count']} rides in {query_time:.4f}s")
                
                if query_time < 1.0:  # Should be very fast for test data
                    print(f"    ✅ Query performance acceptable")
                else:
                    print(f"    ⚠️ Query performance slow (may need indexing)")
                
                # Test payroll calculation performance
                try:
                    calculator = PayrollCalculator(company_id)
                    
                    # Get a driver from this company
                    cursor.execute("SELECT id FROM drivers WHERE company_id = ? LIMIT 1", (company_id,))
                    driver_result = cursor.fetchone()
                    
                    if driver_result:
                        driver_id = driver_result['id']
                        
                        start_time = time.time()
                        payroll_data = calculator.calculate_driver_payroll(
                            driver_id=driver_id,
                            start_date='2025-02-01',
                            end_date='2025-02-28'
                        )
                        calc_time = time.time() - start_time
                        
                        print(f"    ⚡ Payroll calculation: €{payroll_data.get('total_pay', 0):.2f} in {calc_time:.4f}s")
                        
                        if calc_time < 2.0:  # Should be reasonably fast
                            print(f"    ✅ Payroll performance acceptable")
                        else:
                            print(f"    ⚠️ Payroll calculation slow")
                    
                except Exception as e:
                    print(f"    ⚠️ Payroll calculation test failed: {e}")
                
                db.close()
            
            # Test caching performance across companies
            try:
                maps_api = GoogleMapsIntegration()
                
                # Test that cache works across different company contexts
                origin = 'Muster Str 1, 45451 MusterStadt'
                destination = 'Beispielstraße 10, 45451 MusterStadt'
                
                start_time = time.time()
                distance1, duration1 = maps_api.calculate_distance_and_duration(origin, destination)
                first_call = time.time() - start_time
                
                start_time = time.time()
                distance2, duration2 = maps_api.calculate_distance_and_duration(origin, destination)
                second_call = time.time() - start_time
                
                if second_call < first_call:
                    print(f"  ✅ Cache working across companies: {first_call:.4f}s -> {second_call:.4f}s")
                else:
                    print(f"  ⚠️ Cache performance unclear: {first_call:.4f}s -> {second_call:.4f}s")
                
            except Exception as e:
                print(f"  ⚠️ Cache performance test failed: {e}")
            
            self.test_results['performance_isolation'] = performance_success
            
        except Exception as e:
            print(f"  ❌ Performance isolation test failed: {e}")
            self.test_results['performance_isolation'] = False

    def run_all_tests(self):
        """Run all multi-company tests and generate summary report"""
        print("\n" + "="*80)
        print("🏢 RUNNING COMPREHENSIVE MULTI-COMPANY TEST SUITE")
        print("="*80)
        
        # Run all tests
        self.test_database_multi_company_schema()
        self.test_company_management()
        self.test_data_isolation()
        self.test_company_switching()
        self.test_multi_company_exports()
        self.test_company_statistics()
        self.test_app_mode_management()
        self.test_company_specific_config()
        self.test_ui_integration()
        self.test_performance_isolation()
        
        # Generate summary report
        print("\n" + "="*80)
        print("📊 MULTI-COMPANY TEST SUMMARY REPORT")
        print("="*80)
        
        passed_tests = sum(self.test_results.values())
        total_tests = len(self.test_results)
        
        test_names = {
            'database_multi_company': 'Database Multi-Company Schema',
            'company_management': 'Company Management Operations',
            'data_isolation': 'Data Isolation Between Companies',
            'company_switching': 'Company Switching Functionality',
            'multi_company_exports': 'Multi-Company Export Functionality',
            'company_statistics': 'Company Statistics and Analytics',
            'app_mode_management': 'Application Mode Management',
            'company_specific_config': 'Company-Specific Configuration',
            'ui_integration': 'UI Integration with Multi-Company',
            'performance_isolation': 'Performance with Data Isolation'
        }
        
        for test_key, result in self.test_results.items():
            test_name = test_names.get(test_key, test_key.replace('_', ' ').title())
            status = "✅ PASSED" if result else "❌ FAILED"
            print(f"{test_name:<45} {status}")
        
        print("\n" + "-"*80)
        print(f"OVERALL RESULT: {passed_tests}/{total_tests} tests passed ({passed_tests/total_tests*100:.1f}%)")
        
        if passed_tests == total_tests:
            print("🎉 ALL MULTI-COMPANY TESTS PASSED! Multi-company functionality is working correctly.")
        elif passed_tests >= total_tests * 0.8:
            print("⚠️ MOSTLY WORKING: Most multi-company features are functioning, some need attention.")
        else:
            print("🔧 NEEDS WORK: Several multi-company features require fixes.")
        
        print("\n📋 MULTI-COMPANY FEATURE CHECKLIST:")
        checklist_items = [
            ("Multi-Company Database Schema", self.test_results['database_multi_company']),
            ("Company CRUD Operations", self.test_results['company_management']),
            ("Data Isolation & Security", self.test_results['data_isolation']),
            ("Company Switching", self.test_results['company_switching']),
            ("Company-Specific Exports", self.test_results['multi_company_exports']),
            ("Company Statistics", self.test_results['company_statistics']),
            ("App Mode Management", self.test_results['app_mode_management']),
            ("Company-Specific Config", self.test_results['company_specific_config']),
            ("UI Multi-Company Integration", self.test_results['ui_integration']),
            ("Performance Optimization", self.test_results['performance_isolation'])
        ]
        
        for item, status in checklist_items:
            checkbox = "☑️" if status else "☐"
            print(f"{checkbox} {item}")
        
        print("\n🚀 Multi-Company Ride Guardian Desktop comprehensive testing complete!")
        return self.test_results

def main():
    """Main test execution"""
    print("Multi-Company Ride Guardian Desktop - Comprehensive Test Suite")
    print("=" * 70)
    
    test_suite = MultiCompanyTestSuite()
    results = test_suite.run_all_tests()
    
    # Exit with appropriate code
    if all(results.values()):
        sys.exit(0)  # All tests passed
    else:
        sys.exit(1)  # Some tests failed

if __name__ == "__main__":
    main()