#!/usr/bin/env python3
"""
Ride Guardian Desktop - Excel Export Restoration

This script restores and verifies the Excel export functionality:
1. Checks that the enhanced_fahrtenbuch_export.py file is present and working
2. Verifies that all required templates are available
3. Tests export functionality with sample data
4. Fixes common issues with the export system

Run this script to restore the Excel export functionality.
"""

import os
import sys
import importlib
import inspect
import shutil
from datetime import datetime, timedelta
import openpyxl

# Ensure project root in PYTHONPATH
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

# Import necessary modules if available
try:
    from core.database import get_db_connection, get_companies
    
    # Try to import export modules - will fail if they're missing
    try:
        from core.enhanced_fahrtenbuch_export import EnhancedFahrtenbuchExporter
        enhanced_exporter_available = True
    except ImportError:
        enhanced_exporter_available = False
        
    try:
        from core.fahrtenbuch_export import FahrtenbuchExporter
        basic_exporter_available = True
    except ImportError:
        basic_exporter_available = False
        
    # Check if Excel workbook logic is available
    try:
        from core.excel_workbook_logic import ExcelWorkbookLogic
        excel_logic_available = True
    except ImportError:
        excel_logic_available = False
        
except ImportError as e:
    print(f"Error importing required modules: {e}")
    print("Please ensure all required packages are installed.")
    sys.exit(1)

def print_header(text):
    """Print a formatted header"""
    print("\n" + "="*80)
    print(f" {text} ".center(80, "="))
    print("="*80)

def check_template_files():
    """Check if the required Excel template files are present"""
    print_header("Checking Template Files")
    
    template_files = [
        "Fahrtenbuch V6-2.xlsx",
        "Beispiel von 2025-02-01_2025-03-31.xlsx"
    ]
    
    missing_templates = []
    for template in template_files:
        template_path = os.path.join(project_root, template)
        if os.path.exists(template_path):
            file_size = os.path.getsize(template_path) / 1024  # KB
            print(f"✅ Found template: {template} ({file_size:.1f} KB)")
        else:
            print(f"❌ Missing template: {template}")
            missing_templates.append(template)
            
    if missing_templates:
        print("\n⚠️ Some template files are missing. Please obtain these files:")
        for template in missing_templates:
            print(f"  - {template}")
        return False
    else:
        print("✅ All required template files are present")
        return True

def check_exporter_modules():
    """Check if the exporter modules are available and working"""
    print_header("Checking Exporter Modules")
    
    if enhanced_exporter_available:
        print("✅ Enhanced Fahrtenbuch exporter is available")
    else:
        print("❌ Enhanced Fahrtenbuch exporter is missing")
        
    if basic_exporter_available:
        print("✅ Basic Fahrtenbuch exporter is available")
    else:
        print("❌ Basic Fahrtenbuch exporter is missing")
        
    if excel_logic_available:
        print("✅ Excel workbook logic module is available")
    else:
        print("❌ Excel workbook logic module is missing")
        
    # Check if any exporter is available
    if not enhanced_exporter_available and not basic_exporter_available:
        print("\n❌ No Fahrtenbuch exporter modules are available.")
        print("💡 Need to create or restore exporter modules.")
        return False
    else:
        if enhanced_exporter_available:
            # Verify the enhanced exporter has all required methods
            exporter = EnhancedFahrtenbuchExporter(company_id=1)
            required_methods = [
                "export_fahrtenbuch_excel_enhanced",
                "export_stundenzettel_excel",
                "export_fahrtenbuch_pdf"
            ]
            
            missing_methods = [method for method in required_methods 
                             if not hasattr(exporter, method)]
            
            if missing_methods:
                print(f"⚠️ Enhanced exporter is missing methods: {missing_methods}")
                return False
            else:
                print("✅ Enhanced exporter has all required methods")
                return True
        else:
            print("⚠️ Enhanced exporter not available, but basic exporter exists")
            return False

def create_test_data():
    """Create test data for export functionality testing"""
    print_header("Creating Test Data")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if there are any companies
    companies = get_companies()
    if not companies:
        print("❌ No companies found in database")
        print("💡 Creating a test company...")
        
        cursor.execute("""
            INSERT INTO companies (name, headquarters_address) 
            VALUES (?, ?)
        """, ("Muster GmbH", "Muster Str 1, 45451 MusterStadt"))
        company_id = cursor.lastrowid
        conn.commit()
    else:
        company_id = companies[0]['id']
        print(f"✅ Using existing company: {companies[0]['name']}")
    
    # Check if there are any drivers
    cursor.execute("SELECT COUNT(*) FROM drivers WHERE company_id = ?", (company_id,))
    driver_count = cursor.fetchone()[0]
    
    if driver_count == 0:
        print("❌ No drivers found for this company")
        print("💡 Creating a test driver...")
        
        cursor.execute("""
            INSERT INTO drivers (company_id, name, personalnummer) 
            VALUES (?, ?, ?)
        """, (company_id, "Mitarbeiter 1", "(1)"))
        driver_id = cursor.lastrowid
        conn.commit()
    else:
        cursor.execute("SELECT id FROM drivers WHERE company_id = ? LIMIT 1", (company_id,))
        driver_id = cursor.fetchone()[0]
        print(f"✅ Using existing driver (ID: {driver_id})")
    
    # Check if there are any vehicles
    cursor.execute("SELECT COUNT(*) FROM vehicles WHERE company_id = ?", (company_id,))
    vehicle_count = cursor.fetchone()[0]
    
    if vehicle_count == 0:
        print("❌ No vehicles found for this company")
        print("💡 Creating a test vehicle...")
        
        cursor.execute("""
            INSERT INTO vehicles (company_id, make, model, plate_number) 
            VALUES (?, ?, ?, ?)
        """, (company_id, "Toyota", "Corolla", "E-CK12345"))
        conn.commit()
        print("✅ Test vehicle created")
    else:
        print("✅ Vehicles already exist")
    
    # Check if there are any rides
    cursor.execute("SELECT COUNT(*) FROM rides WHERE driver_id = ?", (driver_id,))
    ride_count = cursor.fetchone()[0]
    
    if ride_count < 5:
        print(f"⚠️ Only {ride_count} rides found for test driver")
        print("💡 Creating test rides...")
        
        # Create some test rides for the driver
        headquarters = "Muster Str 1, 45451 MusterStadt"
        
        # Sample addresses
        addresses = [
            "Cecilienstraße 11, 47051 Duisburg",
            "Goetheplatz 4, 45468 Mülheim an der Ruhr",
            "Werftstraße 20, 45468 Mülheim an der Ruhr",
            "Martin-Luther-Straße 5, 46047 Oberhausen",
            "Von-Ossietzky-Ring 45279 Essen"
        ]
        
        # Create rides for the past week
        now = datetime.now()
        for i in range(5):
            pickup_time = now - timedelta(days=i, hours=i)
            dropoff_time = pickup_time + timedelta(hours=1)
            
            # Only create rides if we don't have enough
            if i >= ride_count:
                cursor.execute("""
                    INSERT INTO rides (
                        company_id, driver_id, pickup_time, dropoff_time,
                        pickup_location, destination, 
                        standort_auftragsuebermittlung, abholort, zielort,
                        gefahrene_kilometer, distance_km, duration_minutes,
                        status, vehicle_plate
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    company_id, driver_id, 
                    pickup_time.strftime('%Y-%m-%d %H:%M:%S'),
                    dropoff_time.strftime('%Y-%m-%d %H:%M:%S'),
                    headquarters, addresses[i % len(addresses)],
                    headquarters, headquarters, addresses[i % len(addresses)],
                    10 + i, 10 + i, 15 + i,
                    "Completed", "E-CK12345"
                ))
        
        conn.commit()
        print(f"✅ Created test rides")
    else:
        print(f"✅ {ride_count} rides found for test driver")
    
    conn.close()
    
    return {
        "company_id": company_id,
        "driver_id": driver_id
    }

def test_export_functionality(test_data):
    """Test export functionality"""
    print_header("Testing Export Functionality")
    
    if not enhanced_exporter_available:
        print("❌ Enhanced exporter not available, cannot test export functionality")
        return False
    
    try:
        # Create export directory if it doesn't exist
        export_dir = os.path.join(project_root, "test_exports")
        os.makedirs(export_dir, exist_ok=True)
        
        exporter = EnhancedFahrtenbuchExporter(company_id=test_data["company_id"])
        
        # Test Fahrtenbuch export
        now = datetime.now()
        start_date = (now - timedelta(days=30)).strftime('%Y-%m-%d')
        end_date = now.strftime('%Y-%m-%d')
        
        fahrtenbuch_path = os.path.join(export_dir, "test_fahrtenbuch.xlsx")
        
        print(f"Testing Fahrtenbuch export for driver {test_data['driver_id']}...")
        
        try:
            result = exporter.export_fahrtenbuch_excel_enhanced(
                driver_id=test_data["driver_id"],
                start_date=start_date,
                end_date=end_date,
                output_path=fahrtenbuch_path
            )
            
            if os.path.exists(fahrtenbuch_path):
                file_size = os.path.getsize(fahrtenbuch_path) / 1024  # KB
                print(f"✅ Fahrtenbuch export successful: {fahrtenbuch_path} ({file_size:.1f} KB)")
                
                # Validate the content of the Excel file
                try:
                    wb = openpyxl.load_workbook(fahrtenbuch_path)
                    sheet = wb.active
                    
                    # Check for key headers
                    has_fahrtenbuch_title = "Fahrtenbuch" in sheet['A1'].value if sheet['A1'].value else False
                    
                    if has_fahrtenbuch_title:
                        print("✅ Fahrtenbuch title found in export")
                    else:
                        print("❌ Fahrtenbuch title not found in export")
                    
                    wb.close()
                except Exception as e:
                    print(f"❌ Error validating Excel file: {e}")
            else:
                print(f"❌ Export function ran but file not created: {fahrtenbuch_path}")
                return False
                
        except Exception as e:
            print(f"❌ Fahrtenbuch export failed: {e}")
            return False
        
        # Test Stundenzettel export
        stundenzettel_path = os.path.join(export_dir, "test_stundenzettel.xlsx")
        
        print(f"Testing Stundenzettel export for driver {test_data['driver_id']}...")
        
        try:
            result = exporter.export_stundenzettel_excel(
                driver_id=test_data["driver_id"],
                month=now.month,
                year=now.year,
                output_path=stundenzettel_path
            )
            
            if os.path.exists(stundenzettel_path):
                file_size = os.path.getsize(stundenzettel_path) / 1024  # KB
                print(f"✅ Stundenzettel export successful: {stundenzettel_path} ({file_size:.1f} KB)")
                
                # Validate the content of the Excel file
                try:
                    wb = openpyxl.load_workbook(stundenzettel_path)
                    sheet = wb.active
                    
                    # Check for key headers
                    has_stundenzettel_title = any("Stundenzettel" in str(cell.value) for cell in sheet[1] if cell.value)
                    
                    if has_stundenzettel_title:
                        print("✅ Stundenzettel title found in export")
                    else:
                        print("❌ Stundenzettel title not found in export")
                    
                    wb.close()
                except Exception as e:
                    print(f"❌ Error validating Excel file: {e}")
            else:
                print(f"❌ Export function ran but file not created: {stundenzettel_path}")
                return False
                
        except Exception as e:
            print(f"❌ Stundenzettel export failed: {e}")
            return False
        
        print("✅ All export tests passed successfully")
        return True
        
    except Exception as e:
        print(f"❌ Error testing export functionality: {e}")
        return False

def fix_common_export_issues():
    """Fix common issues with the export system"""
    print_header("Fixing Common Export Issues")
    
    # Check if we need to restore from basic to enhanced exporter
    if not enhanced_exporter_available and basic_exporter_available:
        print("⚠️ Enhanced exporter is missing but basic exporter exists")
        print("💡 Creating enhanced exporter from basic exporter...")
        
        # Check if there's a backup file
        backup_path = os.path.join(project_root, "core", "enhanced_fahrtenbuch_export.py.bak")
        target_path = os.path.join(project_root, "core", "enhanced_fahrtenbuch_export.py")
        
        if os.path.exists(backup_path):
            print(f"✅ Found backup file: {backup_path}")
            print(f"💡 Restoring from backup...")
            
            try:
                shutil.copy(backup_path, target_path)
                print(f"✅ Successfully restored enhanced exporter from backup")
            except Exception as e:
                print(f"❌ Error restoring from backup: {e}")
        else:
            print("❌ No backup file found for enhanced exporter")
            
            # Create a new enhanced exporter based on the basic one
            print("💡 Creating new enhanced exporter based on basic exporter...")
            
            try:
                with open(os.path.join(project_root, "core", "fahrtenbuch_export.py"), "r") as f_basic:
                    basic_content = f_basic.read()
                
                # Modify the content to create an enhanced version
                enhanced_content = basic_content.replace("FahrtenbuchExporter", "EnhancedFahrtenbuchExporter")
                enhanced_content = enhanced_content.replace("class EnhancedFahrtenbuchExporter", 
                                                         "# Enhanced version with template matching\nclass EnhancedFahrtenbuchExporter")
                
                # Add the new export methods
                enhanced_content += """

    def export_fahrtenbuch_excel_enhanced(self, driver_id, start_date, end_date, output_path):
        \"\"\"
        Export Fahrtenbuch with enhanced formatting that matches the template exactly
        \"\"\"
        # This is a placeholder implementation - needs to be replaced with actual template matching code
        return self.export_fahrtenbuch_excel(driver_id, start_date, end_date, output_path)
        
    def export_stundenzettel_excel(self, driver_id, month, year, output_path):
        \"\"\"
        Export Stundenzettel (timesheet) for a driver for a specific month and year
        \"\"\"
        import calendar
        from datetime import datetime
        
        # Calculate start and end dates
        last_day = calendar.monthrange(year, month)[1]
        start_date = f"{year}-{month:02d}-01"
        end_date = f"{year}-{month:02d}-{last_day:02d}"
        
        # Create a new workbook
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Stundenzettel"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        
        # Header
        ws['C1'] = "Stundenzettel"
        ws['C1'].font = Font(bold=True, size=14)
        
        # Green background for employee info
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Get driver and company information
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT d.name, d.personalnummer, c.name as company_name, c.headquarters_address
            FROM drivers d
            JOIN companies c ON d.company_id = c.id
            WHERE d.id = ? AND d.company_id = ?
        ''', (driver_id, self.company_id))
        
        driver_info = cursor.fetchone()
        
        if not driver_info:
            conn.close()
            raise ValueError(f"Driver {driver_id} not found for company {self.company_id}")
        
        # Employee info section with green background
        ws['B3'] = "Mitarbeiter"
        ws['C3'] = driver_info['name']
        ws['C3'].fill = green_fill
        
        ws['B4'] = "Personalnummer"
        ws['C4'] = driver_info['personalnummer']
        ws['C4'].fill = green_fill
        
        ws['F3'] = "Unternehmen:"
        ws['G3'] = driver_info['company_name']
        
        ws['F4'] = ""
        ws['G4'] = driver_info['headquarters_address']
        
        # Table headers
        headers = [
            "Schicht ID", "Tätigkeit", "Datum/Uhrzeit Schichtbeginn", "Datum/Uhrzeit Schichtende",
            "Gesamte Arbeitszeit (Std.)", "Pause (Min.)", "Reale Arbeitszeit (Std.)",
            "Frühschicht (Std.)", "Nachtschicht (Std.)"
        ]
        
        # Add headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Calculate and add shifts
        cursor.execute('''
            SELECT * FROM shifts
            WHERE driver_id = ? AND company_id = ?
            AND substr(datum_uhrzeit_schichtbeginn, 1, 7) = ?
            ORDER BY datum_uhrzeit_schichtbeginn
        ''', (driver_id, self.company_id, f"{year}-{month:02d}"))
        
        shifts = cursor.fetchall()
        
        row = 9
        total_work_hours = 0
        total_early_hours = 0
        total_night_hours = 0
        
        for shift in shifts:
            ws.cell(row=row, column=1, value=shift['schicht_id'] or f"{row-8}-{month}(1)")
            ws.cell(row=row, column=2, value=shift['taetigkeit'] or "Fahren")
            ws.cell(row=row, column=3, value=shift['datum_uhrzeit_schichtbeginn'])
            ws.cell(row=row, column=4, value=shift['datum_uhrzeit_schichtende'])
            ws.cell(row=row, column=5, value=shift['gesamte_arbeitszeit_std'])
            ws.cell(row=row, column=6, value=shift['pause_min'])
            ws.cell(row=row, column=7, value=shift['reale_arbeitszeit_std'])
            ws.cell(row=row, column=8, value=shift['fruehschicht_std'])
            ws.cell(row=row, column=9, value=shift['nachtschicht_std'])
            
            total_work_hours += shift['reale_arbeitszeit_std'] or 0
            total_early_hours += shift['fruehschicht_std'] or 0
            total_night_hours += shift['nachtschicht_std'] or 0
            
            row += 1
        
        # Summary section
        summary_row = row + 2
        
        ws.cell(row=summary_row, column=4, value="Gesamte Arbeitszeit (Monat/Std.)")
        ws.cell(row=summary_row, column=7, value=total_work_hours)
        
        ws.cell(row=summary_row+1, column=4, value="Gesamte Arbeitszeit (Frühschicht, Monat/Std.)")
        ws.cell(row=summary_row+1, column=7, value=total_early_hours)
        
        ws.cell(row=summary_row+2, column=4, value="Gesamte Arbeitszeit (Nachtschicht, Monat/Std.)")
        ws.cell(row=summary_row+2, column=7, value=total_night_hours)
        
        # Get driver's hourly rate
        cursor.execute("SELECT hourly_rate FROM drivers WHERE id = ?", (driver_id,))
        hourly_rate = cursor.fetchone()['hourly_rate'] if cursor.fetchone() else 12.41
        
        ws.cell(row=summary_row, column=8, value="Gesamtbruttolohn")
        ws.cell(row=summary_row, column=9, value=round(total_work_hours * hourly_rate, 2))
        
        ws.cell(row=summary_row+1, column=8, value="Stundenlohn")
        ws.cell(row=summary_row+1, column=9, value=hourly_rate)
        
        conn.close()
        
        # Save the workbook
        wb.save(output_path)
        return output_path
"""
                
                with open(target_path, "w") as f_enhanced:
                    f_enhanced.write(enhanced_content)
                
                print("✅ Created new enhanced exporter")
                print("⚠️ Note: This is a basic implementation that needs further customization")
            except Exception as e:
                print(f"❌ Error creating enhanced exporter: {e}")
    
    # Check if we need to fix issues with an existing enhanced exporter
    elif enhanced_exporter_available:
        exporter = EnhancedFahrtenbuchExporter(company_id=1)
        
        # Check if the export methods are complete
        try:
            source = inspect.getsource(exporter.export_fahrtenbuch_excel_enhanced)
            
            if "placeholder" in source.lower() or len(source.strip().split("\n")) < 10:
                print("⚠️ Enhanced Fahrtenbuch export method appears to be incomplete")
                print("💡 Please review and update the implementation")
            else:
                print("✅ Enhanced Fahrtenbuch export method appears to be complete")
                
        except Exception as e:
            print(f"❌ Error inspecting export method: {e}")
            
    # General advice
    print("\n💡 General fixes for export issues:")
    print("1. Ensure openpyxl and reportlab packages are installed")
    print("2. Check template file paths in the exporter code")
    print("3. Verify that German formatting is used (dd.MM.yyyy, comma for decimal separator)")
    print("4. Make sure headers match exactly: 'Fahrtenbuch', 'Stundenzettel', etc.")
    print("5. Verify that all cells have the correct styles and formatting")

def main():
    """Main restoration function"""
    print_header("Excel Export Restoration")
    print(f"Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Step 1: Check template files
    templates_ok = check_template_files()
    
    # Step 2: Check exporter modules
    exporters_ok = check_exporter_modules()
    
    # Step 3: Create test data
    test_data = create_test_data()
    
    # Step 4: Test export functionality
    if exporters_ok:
        export_ok = test_export_functionality(test_data)
    else:
        export_ok = False
        print("⚠️ Skipping export test as exporters have issues")
    
    # Step 5: Fix common issues
    fix_common_export_issues()
    
    # Summary
    print_header("Restoration Summary")
    print(f"Template Files: {'✅ Available' if templates_ok else '❌ Missing'}")
    print(f"Exporter Modules: {'✅ Working' if exporters_ok else '❌ Issues Detected'}")
    print(f"Export Functionality: {'✅ Working' if export_ok else '❌ Issues Detected'}")
    
    # Next steps
    print_header("Next Steps")
    if templates_ok and exporters_ok and export_ok:
        print("🎉 Excel export functionality has been verified and is working!")
        print("\nTo test the restored functionality:")
        print("1. Run the application: python main.py")
        print("2. Navigate to a view with export options")
        print("3. Export data and verify the Excel file matches the required template")
    else:
        print("⚠️ Some issues remain with the Excel export functionality.")
        print("\nManual fixes required:")
        
        if not templates_ok:
            print("- Obtain the missing template files")
            print("  The exact layouts are critical for compliance")
            
        if not exporters_ok:
            print("- Fix or recreate the enhanced_fahrtenbuch_export.py file")
            print("  Ensure it has all the required export methods")
            
        if not export_ok and exporters_ok:
            print("- Debug the export functionality issues")
            print("  Check for exceptions during the export process")

if __name__ == "__main__":
    main() 