import os
import sys

# If this script is run directly, add the project root to sys.path
# This allows "from core..." imports to work.
if __package__ is None or __package__ == '':
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, '..'))
    if PROJECT_ROOT not in sys.path:
        sys.path.insert(0, PROJECT_ROOT)

"""
Fahrtenbuch Export Module - German Logbook Export to Excel and PDF
Replicates the exact layout and formatting from the provided German templates
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import mm, inch
from datetime import datetime, timedelta
import os
from pathlib import Path # Added for robust path handling
from typing import List, Dict, Optional, Tuple
from core.database import get_db_connection, get_company_config
from core.google_maps import GoogleMapsIntegration

class FahrtenbuchExporter:
    """
    Export German Fahrtenbuch (driving logbook) to Excel and PDF formats
    Matches the exact layout from the provided templates
    """
    
    def __init__(self, company_id: int = 1):
        self.company_id = company_id
        self.db_conn = get_db_connection()
        self.maps_api = GoogleMapsIntegration()
        
        # Load company configuration
        self.company_name = get_company_config(company_id, 'company_name') or 'Muster GmbH'
        self.company_address = get_company_config(company_id, 'headquarters_address') or 'Muster Str 1, 45451 MusterStadt'
        
        # Define and create the default export directory on the Desktop
        self.export_dir = Path.home() / "Desktop" / "RideGuardianExports"
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _get_full_export_path(self, filename: str) -> str:
        """Constructs the full path for the export file in the designated export directory."""
        return str(self.export_dir / filename)
        
    def export_fahrtenbuch_excel(self, driver_id: Optional[int] = None, 
                                start_date: str = None, end_date: str = None,
                                output_path: str = None) -> str:
        """
        Export Fahrtenbuch to Excel format matching the German template
        
        Args:
            driver_id: Specific driver ID or None for all drivers
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format  
            output_path: Custom output file path
            
        Returns:
            Path to the exported Excel file
        """
        
        # Get ride data
        rides_data = self._get_rides_data(driver_id, start_date, end_date)
        
        if not rides_data:
            raise ValueError("Keine Fahrten für den angegebenen Zeitraum gefunden")
            
        # Generate output filename if not provided
        filename_only = ""
        if not output_path:
            date_suffix = f"{start_date}_{end_date}" if start_date and end_date else datetime.now().strftime("%Y-%m-%d")
            driver_suffix = f"_Fahrer_{driver_id}" if driver_id else ""
            filename_only = f"Fahrtenbuch_{date_suffix}{driver_suffix}.xlsx"
            output_path = self._get_full_export_path(filename_only)
        else: # If output_path is provided, use it as is (maybe it's a specific location user chose)
            pass # No change to output_path if it was explicitly given
            
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Group rides by driver and shift
        grouped_data = self._group_rides_by_driver_and_shift(rides_data)
        
        for i, (driver_info_key, driver_data_value) in enumerate(grouped_data.items()):
            # Create worksheet for each driver
            if i == 0:
                ws = wb.active
                ws.title = f"Fahrtenbuch_{driver_data_value['name'][:10]}"
            else:
                ws = wb.create_sheet(f"Fahrtenbuch_{driver_data_value['name'][:10]}")
                
            self._create_fahrtenbuch_excel_sheet(ws, driver_data_value, driver_data_value['shifts'])
            
        # Save workbook
        wb.save(output_path)
        return output_path
        
    def export_fahrtenbuch_pdf(self, driver_id: Optional[int] = None,
                              start_date: str = None, end_date: str = None,
                              output_path: str = None) -> str:
        """
        Export Fahrtenbuch to PDF format matching the German template
        """
        
        # Get ride data
        rides_data = self._get_rides_data(driver_id, start_date, end_date)
        
        if not rides_data:
            raise ValueError("Keine Fahrten für den angegebenen Zeitraum gefunden")
            
        # Generate output filename if not provided
        filename_only = ""
        if not output_path:
            date_suffix = f"{start_date}_{end_date}" if start_date and end_date else datetime.now().strftime("%Y-%m-%d")
            driver_suffix = f"_Fahrer_{driver_id}" if driver_id else ""
            filename_only = f"Fahrtenbuch_{date_suffix}{driver_suffix}.pdf"
            output_path = self._get_full_export_path(filename_only)
        else: # If output_path is provided, use it as is
            pass
            
        # Create PDF document
        doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        # Group rides by driver and shift
        grouped_data = self._group_rides_by_driver_and_shift(rides_data)
        
        story = []
        
        for i, (driver_info_key, driver_data_value) in enumerate(grouped_data.items()):
            if i > 0:
                story.append(PageBreak())
                
            story.extend(self._create_fahrtenbuch_pdf_content(driver_data_value, driver_data_value['shifts']))
            
        # Build PDF
        doc.build(story)
        return output_path
        
    def export_stundenzettel_excel(self, driver_id: int, month: str, year: int,
                                  output_path: str = None) -> str:
        """
        Export Stundenzettel (timesheet) to Excel format matching the German template
        """
        
        # Get shift data for the month
        shifts_data = self._get_shifts_data(driver_id, month, year)
        
        if not shifts_data:
            raise ValueError("Keine Schichtdaten für den angegebenen Monat gefunden")
            
        # Generate output filename if not provided
        filename_only = ""
        if not output_path:
            filename_only = f"Stundenzettel_{year}_{int(month):02d}_Fahrer_{driver_id}.xlsx"
            output_path = self._get_full_export_path(filename_only)
        else: # If output_path is provided, use it as is
            pass
            
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stundenzettel"
        
        self._create_stundenzettel_excel_sheet(ws, shifts_data)
        
        # Save workbook
        wb.save(output_path)
        return output_path
        
    def _get_rides_data(self, driver_id: Optional[int], start_date: str, end_date: str) -> List[Dict]:
        """Get rides data from database with German field names"""
        
        cursor = self.db_conn.cursor()
        
        # Build query
        query = """
            SELECT 
                r.id, r.pickup_time, r.dropoff_time, r.pickup_location, r.destination,
                r.standort_auftragsuebermittlung, r.abholort, r.zielort, 
                r.gefahrene_kilometer, r.verbrauch_liter, r.kosten_euro, r.reisezweck,
                r.is_reserved, r.vehicle_plate, r.shift_id, r.status,
                d.name as fahrer_name, d.personalnummer,
                s.schicht_id, s.start_time as schicht_start, s.end_time as schicht_end
            FROM rides r
            LEFT JOIN drivers d ON r.driver_id = d.id
            LEFT JOIN shifts s ON r.shift_id = s.id
            WHERE r.company_id = ?
        """
        
        params = [self.company_id]
        
        if driver_id:
            query += " AND r.driver_id = ?"
            params.append(driver_id)
            
        if start_date:
            query += " AND DATE(r.pickup_time) >= ?"
            params.append(start_date)
            
        if end_date:
            query += " AND DATE(r.pickup_time) <= ?"
            params.append(end_date)
            
        query += " ORDER BY d.name, s.shift_date, r.pickup_time"
        
        cursor.execute(query, params)
        
        # Get column names from cursor.description
        column_names = [desc[0] for desc in cursor.description]
        
        # Fetch all rows and convert them to dictionaries
        rides = [dict(zip(column_names, row)) for row in cursor.fetchall()]
        
        return rides
        
    def _get_shifts_data(self, driver_id: int, month: str, year: int) -> List[Dict]:
        """Get shifts data for Stundenzettel"""
        
        cursor = self.db_conn.cursor()
        
        query = """
            SELECT 
                s.*, d.name as fahrer_name, d.personalnummer
            FROM shifts s
            LEFT JOIN drivers d ON s.driver_id = d.id
            WHERE s.company_id = ? AND s.driver_id = ?
                AND strftime('%Y', s.shift_date) = ?
                AND strftime('%m', s.shift_date) = ?
            ORDER BY s.shift_date, s.start_time
        """
        
        cursor.execute(query, [self.company_id, driver_id, str(year), f"{int(month):02d}"])
        
        # Get column names from cursor.description
        column_names = [desc[0] for desc in cursor.description]
        
        # Fetch all rows and convert them to dictionaries
        shifts = [dict(zip(column_names, row)) for row in cursor.fetchall()]
        
        return shifts
        
    def _group_rides_by_driver_and_shift(self, rides_data: List[Dict]) -> Dict:
        """Group rides by driver and then by shift"""
        
        grouped = {}
        
        for ride in rides_data:
            driver_key = (ride['fahrer_name'], ride['personalnummer'])
            
            if driver_key not in grouped:
                grouped[driver_key] = {
                    'name': ride['fahrer_name'],
                    'personalnummer': ride['personalnummer'],
                    'shifts': {}
                }
                
            shift_key = ride['schicht_id'] or 'ohne_schicht'
            
            if shift_key not in grouped[driver_key]['shifts']:
                grouped[driver_key]['shifts'][shift_key] = {
                    'schicht_info': {
                        'id': ride['schicht_id'],
                        'start': ride['schicht_start'],
                        'end': ride['schicht_end']
                    },
                    'rides': []
                }
                
            grouped[driver_key]['shifts'][shift_key]['rides'].append(ride)
            
        return grouped
        
    def _create_fahrtenbuch_excel_sheet(self, ws, driver_info: Dict, shifts: Dict):
        """Create Excel sheet matching the German Fahrtenbuch template"""
        
        # Set column widths for 13 columns
        column_widths = [
            ('A', 12), ('B', 10), ('C', 30), ('D', 8), ('E', 22), ('F', 22), # Shortened C, E, F slightly
            ('G', 12), ('H', 10), ('I', 10), ('J', 12), ('K', 10), ('L', 12), ('M', 15) # Added H, I, adjusted G, J, K, L, M
        ]
        
        for col, width in column_widths:
            ws.column_dimensions[col].width = width
            
        # Header section
        # "Fahrtenbuch" title spanning A1 to J1
        ws.merge_cells('A1:J1') 
        ws['A1'] = "Fahrtenbuch"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date range spanning K1 to M1
        ws.merge_cells('K1:M1') 
        first_ride_date_str = "N/A"
        last_ride_date_str = "N/A"

        all_rides_in_sheet = [ride for shift_data in shifts.values() for ride in shift_data['rides']]
        if all_rides_in_sheet:
            pickup_times = [datetime.fromisoformat(r['pickup_time']) for r in all_rides_in_sheet if r['pickup_time']]
            dropoff_times = [datetime.fromisoformat(r['dropoff_time']) for r in all_rides_in_sheet if r['dropoff_time']]
            
            if pickup_times:
                first_ride_time = min(pickup_times)
                first_ride_date_str = first_ride_time.strftime('%d/%m/%Y')
            if dropoff_times: 
                last_ride_time = max(dropoff_times)
                last_ride_date_str = last_ride_time.strftime('%d/%m/%Y')

        ws['K1'] = f"{first_ride_date_str} - {last_ride_date_str}"
        ws['K1'].alignment = Alignment(horizontal='right', vertical='center')
        ws['K1'].font = Font(size=10)

        # Company information section - spans adjusted for 13 columns
        ws.merge_cells('A3:D3')
        ws['A3'] = "Fahrzeug"
        ws['A3'].font = Font(bold=True)
        
        vehicle_name_placeholder = "FEHLEND: Fahrzeugmodell" 
        if all_rides_in_sheet and all_rides_in_sheet[0].get('vehicle_name'):
            vehicle_name_placeholder = all_rides_in_sheet[0]['vehicle_name']
        elif all_rides_in_sheet and all_rides_in_sheet[0].get('vehicle_plate'):
            vehicle_name_placeholder = f"Fzg. mit KN: {all_rides_in_sheet[0]['vehicle_plate']}"
        ws.merge_cells('E3:H3') # E to H
        ws['E3'] = vehicle_name_placeholder
        
        ws.merge_cells('J3:M3') # J to M
        ws['J3'] = "Betriebssitz des Unternehmens:"
        ws['J3'].font = Font(bold=True)
        
        ws.merge_cells('A4:D4')
        ws['A4'] = "Kennzeichen"
        ws['A4'].font = Font(bold=True)
        
        vehicle_plate_placeholder = "FEHLEND: Kennzeichen"
        if all_rides_in_sheet and all_rides_in_sheet[0].get('vehicle_plate'):
            vehicle_plate_placeholder = all_rides_in_sheet[0]['vehicle_plate']
        ws.merge_cells('E4:H4') # E to H
        ws['E4'] = vehicle_plate_placeholder
        
        ws.merge_cells('J4:M4') # J to M
        ws['J4'] = self.company_name
        
        ws.merge_cells('A5:D5')
        ws['A5'] = "Fahrer"
        ws['A5'].font = Font(bold=True)
        
        ws.merge_cells('E5:H5') # E to H
        ws['E5'] = driver_info['name']
        
        ws.merge_cells('J5:M5') # J to M
        ws['J5'] = self.company_address
        
        ws.merge_cells('A6:D6')
        ws['A6'] = "Personalnummer"
        ws['A6'].font = Font(bold=True)
        
        ws.merge_cells('E6:H6') # E to H
        ws['E6'] = driver_info['personalnummer'] or ""
        
        # Table headers (row 8) - 13 columns
        headers = [
            "Datum\nFahrtbeginn", "Uhrzeit\nFahrtbeginn", 
            "Standort des Fahrzeugs bei Auftragsübermittlung", "Ist\nReserve",
            "Abholort", "Zielort", "gefahrene gesamt\nKilometer",
            "Verbrauch (L)", "Kosten (€)", # New Columns
            "Datum\nFahrtende", "Uhrzeit\nFahrtende", "Fahrstatus", "Kennzeichen"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
            
        # Data rows
        current_row = 9
        
        for shift_id, shift_data in shifts.items():
            for ride in shift_data['rides']:
                # Recalculate ride-specific values using business logic
                recalculated_ride_data = self.calculate_business_logic(ride.copy()) # Use a copy
                
                pickup_time = datetime.fromisoformat(ride['pickup_time']) if ride['pickup_time'] else None
                dropoff_time = datetime.fromisoformat(ride['dropoff_time']) if ride['dropoff_time'] else None
                
                # Date and time formatting
                pickup_date = pickup_time.strftime('%d/%m/%Y') if pickup_time else ""
                pickup_time_str = pickup_time.strftime('%H:%M:%S') if pickup_time else ""
                dropoff_date = dropoff_time.strftime('%d/%m/%Y') if dropoff_time else ""
                dropoff_time_str = dropoff_time.strftime('%H:%M:%S') if dropoff_time else ""
                
                row_data = [
                    pickup_date,  # Datum Fahrtbeginn
                    pickup_time_str,  # Uhrzeit Fahrtbeginn
                    ride['standort_auftragsuebermittlung'] or "",  # Standort Auftragsübermittlung
                    "Ja" if ride['is_reserved'] else "Nein",  # Ist Reserve
                    ride['abholort'] or ride['pickup_location'] or "",  # Abholort
                    ride['zielort'] or ride['destination'] or "",  # Zielort
                    f"{recalculated_ride_data.get('gefahrene_kilometer', ride.get('gefahrene_kilometer', 0)):.1f}" if (recalculated_ride_data.get('gefahrene_kilometer') is not None or ride.get('gefahrene_kilometer') is not None) else "",
                    f"{recalculated_ride_data.get('verbrauch_liter', 0):.2f}", # Verbrauch
                    f"{recalculated_ride_data.get('kosten_euro', 0):.2f}", # Kosten
                    dropoff_date,  # Datum Fahrtende
                    dropoff_time_str,  # Uhrzeit Fahrtende
                    ride.get('status', 'Abgeschlossen') or "Abgeschlossen",  # Fahrstatus
                    ride['vehicle_plate'] or ""  # Kennzeichen
                ]
                
                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        top=Side(style='thin'),
                        bottom=Side(style='thin'),
                        left=Side(style='thin'),
                        right=Side(style='thin')
                    )
                    
                current_row += 1
                
        # Add summary section at the bottom
        summary_row_start = current_row + 1 # Start summary a bit lower
        
        # Calculate summary values
        total_km_driver = 0
        total_shift_duration_seconds = 0
        total_pause_seconds = 0
        total_verbrauch_driver = 0.0
        total_kosten_driver = 0.0
        
        all_rides_for_driver = [r for s_data in shifts.values() for r in s_data['rides']]
        recalculated_rides_for_summary = [self.calculate_business_logic(r.copy()) for r in all_rides_for_driver]

        for r_data in recalculated_rides_for_summary: # Use recalculated data for summaries
            total_km_driver += r_data.get('gefahrene_kilometer', 0) or 0
            total_verbrauch_driver += r_data.get('verbrauch_liter', 0) or 0
            total_kosten_driver += r_data.get('kosten_euro', 0) or 0

        processed_shift_ids_for_summary = set()
        for shift_id, shift_data_summary in shifts.items():
            if not shift_id or shift_id == 'ohne_schicht' or shift_id in processed_shift_ids_for_summary:
                continue # Skip rides not associated with a specific shift for pause/work time calc
            
            processed_shift_ids_for_summary.add(shift_id)

            schicht_info = shift_data_summary.get('schicht_info', {})
            schicht_start_str = schicht_info.get('start')
            schicht_end_str = schicht_info.get('end')

            if schicht_start_str and schicht_end_str:
                try:
                    schicht_start_dt = datetime.fromisoformat(schicht_start_str)
                    schicht_end_dt = datetime.fromisoformat(schicht_end_str)
                    
                    current_shift_duration = (schicht_end_dt - schicht_start_dt).total_seconds()
                    if current_shift_duration < 0: current_shift_duration = 0 # Should not happen
                    total_shift_duration_seconds += current_shift_duration
                    
                    # Calculate total ride duration within this specific shift
                    ride_duration_in_shift_seconds = 0
                    for ride_in_shift in shift_data_summary.get('rides', []):
                        pickup_str = ride_in_shift.get('pickup_time')
                        dropoff_str = ride_in_shift.get('dropoff_time')
                        if pickup_str and dropoff_str:
                            try:
                                ride_pickup_dt = datetime.fromisoformat(pickup_str)
                                ride_dropoff_dt = datetime.fromisoformat(dropoff_str)
                                ride_duration_in_shift_seconds += max(0, (ride_dropoff_dt - ride_pickup_dt).total_seconds())
                            except ValueError:
                                pass # Issue parsing ride times

                    current_shift_pause = current_shift_duration - ride_duration_in_shift_seconds
                    if current_shift_pause > 0:
                         total_pause_seconds += current_shift_pause
                except ValueError:
                    pass # Issue parsing shift times

        # Format durations to HH:MM
        total_work_hours_str = f"{int(total_shift_duration_seconds // 3600):02d}:{int((total_shift_duration_seconds % 3600) // 60):02d}"
        total_pause_hours_str = f"{int(total_pause_seconds // 3600):02d}:{int((total_pause_seconds % 3600) // 60):02d}"

        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center')

        # Row 1 of Summary
        ws.merge_cells(f'A{summary_row_start}:D{summary_row_start}')
        ws[f'A{summary_row_start}'] = "Gesamte gefahrene KM:"
        ws[f'A{summary_row_start}'].font = bold_font
        
        ws.merge_cells(f'E{summary_row_start}:G{summary_row_start}')
        ws[f'E{summary_row_start}'] = f"{total_km_driver:.1f} km"
        ws[f'E{summary_row_start}'].alignment = center_align
        
        # Row 2 of Summary
        summary_row_start += 1
        ws.merge_cells(f'A{summary_row_start}:D{summary_row_start}')
        ws[f'A{summary_row_start}'] = "Gesamte Arbeitszeit (Schicht):"
        ws[f'A{summary_row_start}'].font = bold_font

        ws.merge_cells(f'E{summary_row_start}:G{summary_row_start}')
        ws[f'E{summary_row_start}'] = total_work_hours_str
        ws[f'E{summary_row_start}'].alignment = center_align

        # Row 3 of Summary - Pause
        summary_row_start +=1
        ws.merge_cells(f'A{summary_row_start}:D{summary_row_start}')
        ws[f'A{summary_row_start}'] = "Pause Gesamt (Std.):"
        ws[f'A{summary_row_start}'].font = bold_font
        
        ws.merge_cells(f'E{summary_row_start}:G{summary_row_start}')
        ws[f'E{summary_row_start}'] = total_pause_hours_str
        ws[f'E{summary_row_start}'].alignment = center_align

        # Row 4 of Summary - Gesamtverbrauch
        summary_row_start +=1
        ws.merge_cells(f'A{summary_row_start}:D{summary_row_start}')
        ws[f'A{summary_row_start}'] = "Gesamtverbrauch (L):"
        ws[f'A{summary_row_start}'].font = bold_font
        
        ws.merge_cells(f'E{summary_row_start}:G{summary_row_start}')
        ws[f'E{summary_row_start}'] = f"{total_verbrauch_driver:.2f} L"
        ws[f'E{summary_row_start}'].alignment = center_align

        # Row 5 of Summary - Gesamtkosten
        summary_row_start +=1
        ws.merge_cells(f'A{summary_row_start}:D{summary_row_start}')
        ws[f'A{summary_row_start}'] = "Gesamtkosten (€):"
        ws[f'A{summary_row_start}'].font = bold_font
        
        ws.merge_cells(f'E{summary_row_start}:G{summary_row_start}')
        ws[f'E{summary_row_start}'] = f"{total_kosten_driver:.2f} €"
        ws[f'E{summary_row_start}'].alignment = center_align
        
        # Row 6 - Notizen (adjusted row)
        summary_row_start += 1
        ws.merge_cells(f'A{summary_row_start}:M{summary_row_start}') # Span all 13 columns for notes
        ws[f'A{summary_row_start}'] = "Notizen:"
        ws[f'A{summary_row_start}'].font = bold_font
        
    def _create_fahrtenbuch_pdf_content(self, driver_info: Dict, shifts: Dict) -> List:
        """Create PDF content matching the German Fahrtenbuch template"""
        
        story = []
        styles = getSampleStyleSheet()
        
        # Header
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=10*mm, # Adjusted space
            alignment=1  # Center alignment
        )
        
        story.append(Paragraph("Fahrtenbuch", title_style))

        all_rides_in_sheet = [ride for shift_data in shifts.values() for ride in shift_data['rides']]
        vehicle_name_val = "FEHLEND: Fahrzeugmodell"
        vehicle_plate_val = "FEHLEND: Kennzeichen"

        if all_rides_in_sheet:
            first_ride = all_rides_in_sheet[0]
            vehicle_plate_val = first_ride.get('vehicle_plate', 'FEHLEND: Kennzeichen')
            # Assuming vehicle_name might be part of ride data in future
            vehicle_name_val = first_ride.get('vehicle_name', f"Fzg. mit KN: {vehicle_plate_val}")

        
        # Company information table
        company_info_data = [
            [Paragraph("<b>Fahrzeug</b>", styles['Normal']), vehicle_name_val, Paragraph("<b>Betriebssitz des Unternehmens:</b>", styles['Normal'])],
            [Paragraph("<b>Kennzeichen</b>", styles['Normal']), vehicle_plate_val, self.company_name],
            [Paragraph("<b>Fahrer</b>", styles['Normal']), driver_info['name'], self.company_address],
            [Paragraph("<b>Personalnummer</b>", styles['Normal']), driver_info['personalnummer'] or "", ""]
        ]
        
        company_info_table = Table(company_info_data, colWidths=[1.5*inch, 2.5*inch, 3*inch]) # Adjusted colWidths
        company_info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            # ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), # Using Paragraph bold instead
            # ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'), # Using Paragraph bold instead
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey), # Lighter grid
            ('BOTTOMPADDING', (0,0), (-1,-1), 3*mm),
            ('TOPPADDING', (0,0), (-1,-1), 2*mm),
        ]))
        
        story.append(company_info_table)
        story.append(Spacer(1, 8*mm)) # Adjusted spacer
        
        # Main data table headers
        # Using Paragraphs for headers to allow text wrapping and bolding
        header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, alignment=1)
        
        table_headers = [
            Paragraph("Datum<br/>Fahrtbeginn", header_style), Paragraph("Uhrzeit<br/>Fahrtbeginn", header_style),
            Paragraph("Standort des Fahrzeugs bei Auftragsübermittlung", header_style), Paragraph("Ist<br/>Reserve", header_style),
            Paragraph("Abholort", header_style), Paragraph("Zielort", header_style), Paragraph("gefahrene gesamt<br/>Kilometer", header_style),
            Paragraph("Verbrauch<br/>(L)", header_style), Paragraph("Kosten<br/>(€)", header_style), # New Columns
            Paragraph("Datum<br/>Fahrtende", header_style), Paragraph("Uhrzeit<br/>Fahrtende", header_style), Paragraph("Fahrstatus", header_style), Paragraph("Kennzeichen", header_style)
        ]
        
        table_data = [table_headers]
        
        # Data cell style
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=7.5, alignment=1)
        
        for shift_id, shift_data in shifts.items():
            for ride in shift_data['rides']:
                # Recalculate ride-specific values using business logic
                recalculated_ride_data = self.calculate_business_logic(ride.copy()) # Use a copy

                pickup_time = datetime.fromisoformat(ride['pickup_time']) if ride['pickup_time'] else None
                dropoff_time = datetime.fromisoformat(ride['dropoff_time']) if ride['dropoff_time'] else None
                
                row_str_data = [
                    pickup_time.strftime('%d.%m.%Y') if pickup_time else "",
                    pickup_time.strftime('%H:%M') if pickup_time else "",
                    ride['standort_auftragsuebermittlung'] or "",
                    "Ja" if ride['is_reserved'] else "Nein",
                    ride['abholort'] or ride['pickup_location'] or "",
                    ride['zielort'] or ride['destination'] or "",
                    f"{recalculated_ride_data.get('gefahrene_kilometer', ride.get('gefahrene_kilometer', 0)):.1f}" if (recalculated_ride_data.get('gefahrene_kilometer') is not None or ride.get('gefahrene_kilometer') is not None) else "",
                    f"{recalculated_ride_data.get('verbrauch_liter', 0):.2f}", # Verbrauch
                    f"{recalculated_ride_data.get('kosten_euro', 0):.2f}", # Kosten
                    dropoff_time.strftime('%d.%m.%Y') if dropoff_time else "",
                    dropoff_time.strftime('%H:%M') if dropoff_time else "",
                    ride.get('status', 'Abgeschlossen') or "Abgeschlossen",
                    ride['vehicle_plate'] or ""
                ]
                # Wrap cell data in Paragraphs
                table_data.append([Paragraph(str(cell_val), cell_style) for cell_val in row_str_data])
                
        # Create table with appropriate column widths for 13 columns
        col_widths = [
            0.7*inch, 0.6*inch, 1.2*inch, 0.5*inch, 1.0*inch, # Shortened some for new cols
            1.0*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.7*inch, 
            0.6*inch, 0.7*inch, 0.8*inch 
        ] 
        
        main_table = Table(table_data, colWidths=col_widths, repeatRows=1) # Repeat headers
        main_table.setStyle(TableStyle([
            # ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), # Handled by Paragraph style
            # ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'), # Handled by Paragraph style
            # ('FONTSIZE', (0, 0), (-1, -1), 8), # Handled by Paragraph style
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 1*mm), # Added padding
            ('RIGHTPADDING', (0,0), (-1,-1), 1*mm), # Added padding
        ]))
        
        story.append(main_table)
        
        return story
        
    def _create_stundenzettel_excel_sheet(self, ws, shifts_data: List[Dict]):
        """Create Excel sheet for Stundenzettel (timesheet) matching the German template"""
        
        # This would implement the timesheet layout from the provided template
        # Similar structure to Fahrtenbuch but with shift-specific data
        pass
        
    def calculate_business_logic(self, ride: Dict) -> Dict:
        """
        Calculate business logic values like fuel consumption, costs, etc.
        Replicates Excel formulas from the reference workbook.
        Modifies the ride dictionary in place (or returns a new one with calculated values).
        """
        
        distance = ride.get('gefahrene_kilometer', 0) or 0.0 # Ensure float
        
        # Get fuel consumption rate from config (ensure they are floats)
        try:
            fuel_consumption_per_100km = float(get_company_config(self.company_id, 'default_fuel_consumption') or 8.5)
        except (ValueError, TypeError):
            fuel_consumption_per_100km = 8.5
        
        try:
            fuel_cost_per_liter = float(get_company_config(self.company_id, 'fuel_cost_per_liter') or 1.45)
        except (ValueError, TypeError):
            fuel_cost_per_liter = 1.45
        
        # Calculate fuel consumption
        if distance > 0 and fuel_consumption_per_100km > 0:
            calculated_fuel_consumption = (distance * fuel_consumption_per_100km) / 100.0
        else:
            calculated_fuel_consumption = 0.0
        
        ride['verbrauch_liter'] = calculated_fuel_consumption # Update/add verbrauch_liter
        
        # Calculate travel cost
        if calculated_fuel_consumption > 0 and fuel_cost_per_liter > 0:
            calculated_cost_euro = calculated_fuel_consumption * fuel_cost_per_liter
        else:
            calculated_cost_euro = 0.0
            
        ride['kosten_euro'] = calculated_cost_euro # Update/add kosten_euro
        
        # Placeholder for other calculations based on Excel logic if needed
        # e.g., ride['some_other_value'] = ...

        return ride # Return the modified ride dictionary

if __name__ == '__main__':
    # Note: sys.path is already adjusted by the block at the top of this file.
    # We only need to import modules specific to the test execution here if not already global.
    from datetime import datetime, timedelta # Already global but good to be explicit for test context
    from core.database import initialize_database # This import is specific to the test setup

    print("Running FahrtenbuchExporter for testing (with top-level path fix)...")

    # It's a good practice to initialize the database if your tests rely on it
    # Ensure this is safe for your test environment (e.g., uses a test DB or is idempotent)
    try:
        print("Initializing database for test...")
        initialize_database() # Make sure this function exists and is correctly imported
        print("Database initialized.")
    except Exception as e:
        print(f"Error initializing database: {e}")
        print("Proceeding with test, but DB operations might fail if schema is missing.")

    # Example: Test exporting an Excel file for a default company (ID 1)
    # for the last 7 days, for all drivers.
    try:
        exporter = FahrtenbuchExporter(company_id=1) # Assuming company_id 1 exists for testing
        
        end_date_dt = datetime.now()
        start_date_dt = end_date_dt - timedelta(days=7)
        
        start_date_str = start_date_dt.strftime("%Y-%m-%d")
        end_date_str = end_date_dt.strftime("%Y-%m-%d")

        print(f"Attempting to export Excel for company 1, from {start_date_str} to {end_date_str}")
        
        # You might want to specify a dedicated test output path
        # For now, it will use the default ~/Desktop/RideGuardianExports/
        # test_output_path = os.path.join(PROJECT_ROOT, "test_fahrtenbuch.xlsx") 
        
        exported_file = exporter.export_fahrtenbuch_excel(
            driver_id=None, # For all drivers
            start_date=start_date_str,
            end_date=end_date_str
            # output_path=test_output_path # Uncomment to use a specific test path
        )
        print(f"Test Excel export successful: {exported_file}")

    except ValueError as ve:
        print(f"Test export failed (ValueError): {ve}")
        print("This might be normal if no ride data is found for the test period.")
    except Exception as e:
        print(f"An unexpected error occurred during test export: {e}")
        import traceback
        traceback.print_exc()